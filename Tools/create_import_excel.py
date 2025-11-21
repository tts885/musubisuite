#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Dataverse用OCRテーブル定義とサンプルデータをExcel形式で出力
Power Apps環境にインポート可能な形式
"""

import os
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_dataverse_excel():
    """DataverseインポートExcel作成"""
    
    wb = Workbook()
    wb.remove(wb.active)  # デフォルトシートを削除
    
    # ============================================
    # 1. メニューセクションテーブル
    # ============================================
    ws_menu = wb.create_sheet('OCRMenuSections')
    
    # ヘッダー
    menu_headers = [
        'cr_name',  # プライマリ列
        'cr_description',
        'cr_displayorder',
        'cr_isdefault',
        'cr_color',
    ]
    ws_menu.append(menu_headers)
    
    # サンプルデータ
    menu_data = [
        ['すべてのドキュメント', '全ドキュメントを表示', 1, True, '#3b82f6'],
    ]
    
    for row in menu_data:
        ws_menu.append(row)
    
    # スタイル適用
    apply_header_style(ws_menu, len(menu_headers))
    auto_fit_columns(ws_menu)
    
    # ============================================
    # 2. フォルダテーブル
    # ============================================
    ws_folders = wb.create_sheet('OCRFolders')
    
    # ヘッダー
    folder_headers = [
        'cr_name',  # プライマリ列
        'cr_description',
        'cr_color',
        'cr_parentfolderid_lookup',  # 親フォルダ (Lookup)
        'cr_path',
        'cr_menusectionid_lookup',  # メニューセクション (Lookup)
    ]
    ws_folders.append(folder_headers)
    
    # サンプルデータ (階層構造)
    folder_data = [
        # ルートフォルダ
        ['請求書', '取引先からの請求書類', '#3b82f6', '', '/請求書', 'すべてのドキュメント'],
        ['見積書', 'クライアント向け見積書類', '#10b981', '', '/見積書', 'すべてのドキュメント'],
        ['契約書', '契約関連書類', '#8b5cf6', '', '/契約書', 'すべてのドキュメント'],
        # サブフォルダ - 請求書配下
        ['2024年度', '2024年度の請求書', '#3b82f6', '請求書', '/請求書/2024年度', 'すべてのドキュメント'],
        ['2023年度', '2023年度の請求書', '#3b82f6', '請求書', '/請求書/2023年度', 'すべてのドキュメント'],
        # サブフォルダ - 見積書配下
        ['2024年度見積', '2024年度の見積書', '#10b981', '見積書', '/見積書/2024年度見積', 'すべてのドキュメント'],
        # サブフォルダ - 契約書配下
        ['承認待ち', '未承認の契約書', '#f59e0b', '契約書', '/契約書/承認待ち', 'すべてのドキュメント'],
    ]
    
    for row in folder_data:
        ws_folders.append(row)
    
    # スタイル適用
    apply_header_style(ws_folders, len(folder_headers))
    auto_fit_columns(ws_folders)
    
    # ============================================
    # 3. ドキュメントテーブル
    # ============================================
    ws_docs = wb.create_sheet('OCRDocuments')
    
    # ヘッダー
    doc_headers = [
        'cr_filename',  # プライマリ列
        'cr_filetype',
        'cr_filesize',
        'cr_fileurl',
        'cr_uploadedby',
        'cr_tags',
        'cr_folderid_lookup',  # フォルダ (Lookup)
        'cr_status',  # Picklist
    ]
    ws_docs.append(doc_headers)
    
    # サンプルデータ
    doc_data = [
        ['請求書サンプル１.png', 'image/png', 245600, '/uploads/請求書サンプル１.png', 'user_001', '請求書,2024年11月', '2024年度', '完了'],
        ['請求書サンプル２.png', 'image/png', 236789, '/uploads/請求書サンプル２.png', 'user_002', '請求書,2024年10月', '2024年度', '完了'],
        ['1012_請求書サンプル４.png', 'image/png', 212345, '/uploads/1012_請求書サンプル４.png', 'user_002', '請求書,2024年11月', '2024年度', '完了'],
        ['1013_請求書サンプル３.png', 'image/png', 256789, '/uploads/1013_請求書サンプル３.png', 'user_002', '請求書,2024年9月', '2024年度', '完了'],
        ['請求書サンプル５.png', 'image/png', 234567, '/uploads/請求書サンプル５.png', 'user_001', '請求書,2024年11月', '請求書', '完了'],
        ['請求書サンプル６.png', 'image/png', 242890, '/uploads/請求書サンプル６.png', 'user_002', '請求書,2024年11月', '請求書', '完了'],
        ['1010_サンプル請求書３.jpg', 'image/jpeg', 312456, '/uploads/1010_サンプル請求書３.jpg', 'user_001', '請求書,2024年11月', '見積書', '完了'],
        ['1011_サンプル請求書２.png', 'image/png', 267890, '/uploads/1011_サンプル請求書２.png', 'user_002', '請求書,2024年11月', '2024年度見積', '完了'],
        ['請求書_2024_005.pdf', 'application/pdf', 189234, '/uploads/請求書_2024_005.pdf', 'user_001', '請求書,処理待ち', '承認待ち', '処理待ち'],
        ['見積書_2024_128.pdf', 'application/pdf', 201456, '/uploads/見積書_2024_128.pdf', 'user_003', '見積書,処理待ち', '承認待ち', '処理待ち'],
    ]
    
    for row in doc_data:
        ws_docs.append(row)
    
    # スタイル適用
    apply_header_style(ws_docs, len(doc_headers))
    auto_fit_columns(ws_docs)
    
    # ============================================
    # 4. OCR処理結果テーブル
    # ============================================
    ws_results = wb.create_sheet('OCRResults')
    
    # ヘッダー
    result_headers = [
        'cr_name',  # プライマリ列（ドキュメント名_結果）
        'cr_documentid_lookup',  # ドキュメント (Lookup)
        'cr_overallconfidence',
        'cr_processedat',
        'cr_status',  # Picklist
    ]
    ws_results.append(result_headers)
    
    # サンプルデータ
    result_data = [
        ['請求書サンプル１_結果', '請求書サンプル１.png', 0.96, '2024-11-17 10:30:00', '完了'],
        ['請求書サンプル２_結果', '請求書サンプル２.png', 0.95, '2024-11-17 10:35:00', '完了'],
        ['1012_請求書サンプル４_結果', '1012_請求書サンプル４.png', 0.97, '2024-11-17 10:40:00', '完了'],
        ['1013_請求書サンプル３_結果', '1013_請求書サンプル３.png', 0.94, '2024-11-17 10:45:00', '完了'],
        ['請求書サンプル５_結果', '請求書サンプル５.png', 0.93, '2024-11-17 11:00:00', '完了'],
        ['請求書サンプル６_結果', '請求書サンプル６.png', 0.96, '2024-11-17 11:05:00', '完了'],
        ['1010_サンプル請求書３_結果', '1010_サンプル請求書３.jpg', 0.95, '2024-11-17 11:10:00', '完了'],
        ['1011_サンプル請求書２_結果', '1011_サンプル請求書２.png', 0.92, '2024-11-17 11:15:00', '完了'],
    ]
    
    for row in result_data:
        ws_results.append(row)
    
    # スタイル適用
    apply_header_style(ws_results, len(result_headers))
    auto_fit_columns(ws_results)
    
    # ============================================
    # 5. OCRフィールドテーブル
    # ============================================
    ws_fields = wb.create_sheet('OCRFields')
    
    # ヘッダー
    field_headers = [
        'cr_name',  # プライマリ列（ラベル_値）
        'cr_ocrresultid_lookup',  # OCR結果 (Lookup)
        'cr_label',
        'cr_value',
        'cr_confidence',
        'cr_fieldtype',  # Picklist
        'cr_boundingbox_x',
        'cr_boundingbox_y',
        'cr_boundingbox_width',
        'cr_boundingbox_height',
    ]
    ws_fields.append(field_headers)
    
    # サンプルデータ（請求書サンプル１のフィールド）
    field_data = [
        ['請求書番号_INV-2024-001', '請求書サンプル１_結果', '請求書番号', 'INV-2024-001', 0.98, 'テキスト', 450, 80, 200, 30],
        ['発行日_2024-11-15', '請求書サンプル１_結果', '発行日', '2024年11月15日', 0.95, '日付', 450, 120, 180, 30],
        ['請求先_サンプル商事', '請求書サンプル１_結果', '請求先会社名', '株式会社サンプル商事', 0.97, 'テキスト', 100, 200, 300, 35],
        ['合計金額_935000', '請求書サンプル１_結果', '合計金額', '¥935,000', 0.99, '数値', 500, 680, 150, 35],
        # 請求書サンプル２のフィールド
        ['請求書番号_INV-2024-002', '請求書サンプル２_結果', '請求書番号', 'INV-2024-002', 0.96, 'テキスト', 450, 80, 200, 30],
        ['発行日_2024-11-16', '請求書サンプル２_結果', '発行日', '2024年11月16日', 0.94, '日付', 450, 120, 180, 30],
        ['請求先_テクノロジー', '請求書サンプル２_結果', '請求先会社名', '株式会社テクノロジー', 0.95, 'テキスト', 100, 200, 300, 35],
        ['合計金額_1250000', '請求書サンプル２_結果', '合計金額', '¥1,250,000', 0.97, '数値', 500, 680, 150, 35],
    ]
    
    for row in field_data:
        ws_fields.append(row)
    
    # スタイル適用
    apply_header_style(ws_fields, len(field_headers))
    auto_fit_columns(ws_fields)
    
    # ============================================
    # 6. 説明シート
    # ============================================
    ws_readme = wb.create_sheet('README', 0)
    
    readme_content = [
        ['Dataverse OCRテーブル インポート用Excel'],
        [''],
        ['作成日時', datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')],
        [''],
        ['【インポート手順】'],
        ['1. Power Apps Maker Portal (https://make.powerapps.com) にアクセス'],
        ['2. 左メニューから「Tables」を選択'],
        ['3. 各シートごとに以下を実施:'],
        ['   a) 「Import」→「Import data from Excel」を選択'],
        ['   b) 対応するシート名のデータをインポート'],
        ['   c) 列マッピングを確認'],
        [''],
        ['【テーブル一覧】'],
        ['1. OCRMenuSections - メニューセクション (1件)'],
        ['2. OCRFolders - フォルダ階層 (7件)'],
        ['3. OCRDocuments - ドキュメント (10件)'],
        ['4. OCRResults - OCR処理結果 (8件)'],
        ['5. OCRFields - 抽出フィールド (8件)'],
        [''],
        ['【Lookup列について】'],
        ['- "_lookup"サフィックスの列は他テーブルへの参照です'],
        ['- インポート時に参照先レコードの主キー(名前)を指定してください'],
        [''],
        ['【Picklist列について】'],
        ['- cr_status: 「処理待ち」「処理中」「完了」「エラー」'],
        ['- cr_fieldtype: 「テキスト」「数値」「日付」「金額」「選択肢」'],
        [''],
        ['【注意事項】'],
        ['- インポート順序: MenuSections → Folders → Documents → Results → Fields'],
        ['- Lookup列は参照先テーブルのデータが先に存在する必要があります'],
        ['- ファイルURLは実際の環境に合わせて修正してください'],
    ]
    
    for row_idx, row in enumerate(readme_content, start=1):
        if len(row) == 1:
            ws_readme.cell(row=row_idx, column=1, value=row[0])
        else:
            for col_idx, value in enumerate(row, start=1):
                ws_readme.cell(row=row_idx, column=col_idx, value=value)
    
    # READMEスタイル
    title_cell = ws_readme['A1']
    title_cell.font = Font(size=14, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='3b82f6', end_color='3b82f6', fill_type='solid')
    
    ws_readme.column_dimensions['A'].width = 80
    
    # ============================================
    # 保存
    # ============================================
    output_path = os.path.join(
        os.path.dirname(os.path.dirname(__file__)),
        'Dataverse_OCR_Import.xlsx'
    )
    
    wb.save(output_path)
    return output_path


def apply_header_style(ws, num_cols):
    """ヘッダー行にスタイルを適用"""
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='3b82f6', end_color='3b82f6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    border_style = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_style


def auto_fit_columns(ws):
    """列幅を自動調整"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def main():
    """メイン処理"""
    print("=" * 60)
    print("Dataverse インポート用Excel作成")
    print("=" * 60)
    print()
    
    print("Excelファイル作成中...")
    output_path = create_dataverse_excel()
    
    print()
    print(f"✓ Excelファイルを作成しました:")
    print(f"  {output_path}")
    print()
    print("次のステップ:")
    print("  1. Excelファイルを開いてデータを確認")
    print("  2. Power Apps Maker Portalで各テーブルを作成")
    print("  3. テーブル作成後、Excelからデータをインポート")
    print()


if __name__ == "__main__":
    main()
